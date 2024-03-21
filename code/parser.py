
import lzma
import json
from awpy.parser import DemoParser
import pandas as pd
import os, shutil
import patoolib
import re
from openpyxl import load_workbook

# Folder where demos are stored
mainPath = "C:/Users/Matias/Documents/UDESA/Tesis_maestria/Replication files"
pathToDemos = f"{mainPath}/input"
pathToJsons = f"{mainPath}/jsons"
output = f"{mainPath}/output"

os.chdir(mainPath)

def get_first_empty_row(excel_file, sheet_name='Sheet1'):
    if not os.path.exists(excel_file):
        return 1  # File does not exist, start at row 1
    workbook = load_workbook(excel_file, read_only=True)
    if sheet_name not in workbook.sheetnames:
        return 1  # Sheet does not exist, start at row 1
    worksheet = workbook[sheet_name]
    for i, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if all(cell is None for cell in row):
            return i  # Found the first empty row
    return i + 1  # All rows are filled, start at the next row

# Function to read .xz archives from ESTA
def read_parsed_demo(filename):
  with lzma.LZMAFile(filename, "rb") as f:
    d = json.load(f)
    return d

def generate_vector_state(frame, map_name):
    """Returns a game state in a dictionary format.

    Args:
        frame (dict) : Dict output of a frame generated from the DemoParser class
        map_name (string): String indicating the map name

    Returns:
        A dict with keys for each feature.
    """
    game_state = {}
    game_state["mapName"] = map_name
    game_state["secondsSincePhaseStart"] = frame["seconds"]
    game_state["bombPlanted"] = frame["bombPlanted"]
    game_state["bombsite"] = frame["bombsite"]
    game_state["totalSmokes"] = len(frame["smokes"])
    game_state["totalFires"] = len(frame["fires"])

    # Team specific info (CT)
    game_state["ctAlive"] = 0
    game_state["ctHp"] = 0
    game_state["ctArmor"] = 0
    game_state["ctHelmet"] = 0
    game_state["ctEq"] = 0
    game_state["ctUtility"] = 0
    game_state["ctEqValStart"] = 0
    game_state["ctBombZone"] = 0
    game_state["defusers"] = 0
    game_state["ctNone"] = 0
    game_state["ctCash"] = 0
    if frame["ct"]["players"] != None:
        for p in frame["ct"]["players"]:
            game_state["ctEqValStart"] += p["equipmentValueFreezetimeEnd"]
            game_state["ctCash"] += p["cash"]
            if p["isAlive"]:
                game_state["ctAlive"] += 1
                game_state["ctHp"] += p["hp"]
                game_state["ctArmor"] += p["armor"]
                game_state["ctHelmet"] += p["hasHelmet"]
                game_state["ctEq"] += p["equipmentValue"]
                game_state["ctUtility"] += p["totalUtility"]
                game_state["defusers"] += p["hasDefuse"]
                if p["isInBombZone"]:
                    game_state["ctBombZone"] += 1
    else: 
        game_state["ctNone"] = 1

    # Team specific info (T)
    game_state["tAlive"] = 0
    game_state["tHp"] = 0
    game_state["tArmor"] = 0
    game_state["tHelmet"] = 0
    game_state["tEq"] = 0
    game_state["tUtility"] = 0
    game_state["tEqValStart"] = 0
    game_state["tHoldingBomb"] = 0
    game_state["tBombZone"] = 0
    game_state["tNone"] = 0
    game_state["tCash"] = 0
    if frame["t"]["players"] != None:
        for p in frame["t"]["players"]:
            game_state["tEqValStart"] += p["equipmentValueFreezetimeEnd"]
            game_state["tCash"] += p["cash"]
            if p["isAlive"]:
                game_state["tAlive"] += 1
                game_state["tHp"] += p["hp"]
                game_state["tArmor"] += p["armor"]
                game_state["tHelmet"] += p["hasHelmet"]
                game_state["tEq"] += p["equipmentValue"]
                game_state["tUtility"] += p["totalUtility"]
                if p["isInBombZone"]:
                    game_state["tBombZone"] += 1
                if p["hasBomb"]:
                    game_state["tHoldingBomb"] = 1
    else: 
        game_state["tNone"] = 1

    return game_state

def get_frames(gameRoundsDataframe):
    frames = []
    for ronda in gameRoundsDataframe:
        frames.append(ronda["frames"][-1])
        
    return frames

# Main dataframe
round_state_df = pd.DataFrame()

carpetas = []

mapNumberPattern = re.compile(r'm[1-5]')
partNumberPattern = re.compile(r'p[1-5]')

rarFiles = [file for file in os.listdir() if file.endswith('.rar')]

for file in rarFiles:
    
    demoFiles = [file for file in os.listdir() if file.endswith('.dem')]
    
    for demo in demoFiles: #Remove demos from other matches that failed to be moved
        os.remove(demo)
    
    print(f"Extracting file: {file}")
    patoolib.extract_archive("%s" % file, outdir="%s" %pathToDemos, verbosity=0)

    # Grab demo names
    demoFiles = [file for file in os.listdir() if file.endswith('.dem')]

    ### Itero parser over demos
    for demo in demoFiles:
        
        print(f"Parsing {demo}")
        demo_parser = DemoParser(
        demofile = "%s" % demo,
        parse_rate=128, 
        buy_style="hltv",
        parse_chat = True
        )
    
        try:
            # Parse the demofile, output results to dictionary
            df = demo_parser.parse(return_type="json", clean = True)
            
            mapNumberMatch = mapNumberPattern.search(demo)
            partNumberMatch = partNumberPattern.search(demo)
            
            if mapNumberMatch:
                
                newname = file[:-4]  + "-" + mapNumberMatch.group()
                
            else:
                 newname = file[:-4] + "-m1"
                 
            if partNumberMatch:
                newname = newname + "-" + partNumberMatch.group()
            
            os.remove(demo)
            
            print(f"Removed {demo}")
            print(f"Renaming {demo[:-4]}.json to {newname} and moving to {pathToJsons}")
            
            os.rename(f'{demo[:-4]}.json', f'{newname}.json')
            shutil.move(f'{newname}.json', f"{pathToJsons}")
            
        except Exception as e:
            print(f"An error occured!")
            print(e)
            continue
    
    os.remove(file)
    
os.chdir(pathToJsons)

# Main dataframe
round_state_df = pd.DataFrame()

# Grab demo names

jsonFiles = [file for file in os.listdir() if file.endswith(".json")]
demos = []
for demo in jsonFiles:
    
    print(f"{demo}")
    with open(demo, encoding="utf-8") as demo_json:
            df = json.load(demo_json)
                
    # Grab round end frames (last frame of every round)
    frames = []
        
    try:
        for ronda in df["gameRounds"]:
            frames.append(ronda["frames"][-1])
    except Exception as e:
        print(f"An error occured getting round frames: {e}")
        os.remove(demo)
        continue

    mapa = df["mapName"]
    

    # Generate vectors for each frame and pass them to df
    states = []
    for f in frames:
        game_state = generate_vector_state(f, mapa)
        states.append(game_state)
    states = pd.DataFrame(states)
    states["matchID"] = (demo[:-5])
    
    # Get total freeze time
    roundNum = []
    isWarmup = []
    freezeTimeEnd = []
    startTick = []
    ctTeam = []
    tTeam = []
    ctScore = []
    tScore = []
    endCTScore = []
    endTScore = []
    winningSide = []
    for round in df["gameRounds"]:
        freezeTimeEnd.append(round["freezeTimeEndTick"])
        startTick.append(round["startTick"])
        ctTeam.append(round["ctTeam"])
        tTeam.append(round["tTeam"])
        ctScore.append(round["ctScore"])
        tScore.append(round["tScore"])
        endCTScore.append(round["endCTScore"])
        endTScore.append(round["endTScore"])
        winningSide.append(round["winningSide"])
        roundNum.append(round["roundNum"])
        isWarmup.append(round['isWarmup'])

    states["freezeTimeEndTick"] = freezeTimeEnd
    states["startTick"] = startTick
    states["ctTeam"] = ctTeam
    states["tTeam"] = tTeam
    states["ctScore"] = ctScore
    states["tScore"] = tScore
    states["endCTScore"] = endCTScore
    states["endTScore"] = endTScore
    states["winningSide"] = winningSide
    states["freezeTimeTotal"] = states["freezeTimeEndTick"] - states["startTick"] 
    states["pause"] = 0
    states["roundNum"] = roundNum
    states['isWarmup'] = isWarmup
        
    max_freezeTimeEndTick = states["freezeTimeEndTick"].max()
    
    # Get the actual number of rounds in the DataFrame
    num_rounds = len(states)

    # Loop to fill player names for each position
    # for i in range(5): 
        
    #     if num_rounds <= 15:
    #         # If there are 15 or fewer rounds, fill names for all available rounds (this is for incomplete matches)
    #         states.loc[:num_rounds-1, f"ct_p{i+1}"] = df["gameRounds"][1]["ctSide"]["players"][i]["playerName"]
    #         states.loc[:num_rounds-1, f"t_p{i+1}"] = df["gameRounds"][1]["tSide"]["players"][i]["playerName"]
    #     else:
    #         # If there are more than 15 rounds, fill names according to the original plan
    #         states.loc[:15, f"ct_p{i+1}"] = df["gameRounds"][1]["ctSide"]["players"][i]["playerName"]
    #         states.loc[:15, f"t_p{i+1}"] = df["gameRounds"][1]["tSide"]["players"][i]["playerName"]
    #         states.loc[15:num_rounds-1, f"ct_p{i+1}"] = df["gameRounds"][1]["tSide"]["players"][i]["playerName"]
    #         states.loc[15:num_rounds-1, f"t_p{i+1}"] = df["gameRounds"][1]["ctSide"]["players"][i]["playerName"]
    
    # Initialize variables to store the best round index and the maximum counts of players found
    best_round_index = None
    max_ct_players = 0
    max_t_players = 0

    # Iterate through each round to find the one with the most complete information
    for round_index, game_round in enumerate(df["gameRounds"]):
        # Check if ctSide or tSide is not None before trying to access "players"
        ct_players_count = len(game_round["ctSide"]["players"]) if game_round.get("ctSide") and game_round["ctSide"].get("players") else 0
        t_players_count = len(game_round["tSide"]["players"]) if game_round.get("tSide") and game_round["tSide"].get("players") else 0
        
        # Update if this round has more complete information than previous rounds
        if ct_players_count == 5 and t_players_count == 5:
            best_round_index = round_index
            break  # Exit loop if a round with complete info for both teams is found
        elif ct_players_count > max_ct_players or t_players_count > max_t_players:
            max_ct_players, max_t_players = ct_players_count, t_players_count
            best_round_index = round_index


    # Define a function to safely get player names or return an empty string if not possible
    def get_player_name(players, index):
        return players[index]["playerName"] if index < len(players) else ""

    # Use the best found round to fill the DataFrame, if any round was found
    if best_round_index is not None:
        for i in range(5):
            ct_players = df["gameRounds"][best_round_index]["ctSide"]["players"] if df["gameRounds"][best_round_index]["ctSide"] is not None else []
            t_players = df["gameRounds"][best_round_index]["tSide"]["players"] if df["gameRounds"][best_round_index]["tSide"] is not None else []
            ct_player_name = get_player_name(ct_players, i)
            t_player_name = get_player_name(t_players, i)
            
            if num_rounds <= 15:
                states.loc[:num_rounds-1, f"ct_p{i+1}"] = ct_player_name
                states.loc[:num_rounds-1, f"t_p{i+1}"] = t_player_name
            else:
                states.loc[:15, f"ct_p{i+1}"] = ct_player_name
                states.loc[:15, f"t_p{i+1}"] = t_player_name
                states.loc[15:num_rounds-1, f"ct_p{i+1}"] = t_player_name
                states.loc[15:num_rounds-1, f"t_p{i+1}"] = ct_player_name
    else:
        # If no information was found at all, fill all player names with empty strings
        for i in range(5):
            states.loc[:, f"ct_p{i+1}"] = ""
            states.loc[:, f"t_p{i+1}"] = ""

    # Initialize a dictionary to keep track of the message count for each round
    message_counter_per_round = {}
    # Dictionary to keep track of the last inserted message and tick for each round
    last_inserted = {}

    for i, msg in enumerate(df['chatMessages']):
        
        if msg["tick"] <= max_freezeTimeEndTick:
            # Find the index for the current message's round
            index = states.index[states["freezeTimeEndTick"] >= msg["tick"]][0]
            round_number = states.at[index, "roundNum"]

            # Initialize or increment the counter for this round
            if round_number not in message_counter_per_round:
                message_counter_per_round[round_number] = 1
            else:
                # Increment the message counter for existing rounds
                message_counter_per_round[round_number] += 1

            # Prepare column names for this message
            msg_col_suffix = message_counter_per_round[round_number]
            player_col_name = f"player{msg_col_suffix}"
            msg_col_name = f"msg{msg_col_suffix}"
            msg_tick_col_name = f"msgTick{msg_col_suffix}"

            # Skip insertion if this message and tick match the last inserted ones for the same round
            if round_number in last_inserted and last_inserted[round_number] == (msg["text"]):
                # Decrement message counter as this message will not be added
                message_counter_per_round[round_number] -= 1
                continue  # Skip to the next message

            # Ensure columns exist for player, message, and message tick, adding them if necessary
            if msg_col_name not in states.columns:
                states[player_col_name] = ""
                states[msg_col_name] = ""
                states[msg_tick_col_name] = ""

            # Record the message details
            if msg.get("params") and len(msg["params"]) > 0:
                states.at[index, player_col_name] = msg["params"][0]
            states.at[index, msg_col_name] = msg["text"].replace(",", "").replace("'", "").replace('"', '').strip()
            states.at[index, msg_tick_col_name] = msg["tick"]

            # Update last inserted message and tick for this round
            last_inserted[round_number] = (msg["text"])

            # Check for specific keywords in message to set 'pause'
            if any(keyword in msg["text"] for keyword in ["tech", "pause", "A player disconnected, auto pausing.", "Waiting for both teams and admin to ready to continue."]):
                states.at[index, "pause"] = 1
    
    round_state_df = pd.concat([round_state_df, pd.DataFrame(states)])
    
    os.remove(demo)
        
round_state_df.reset_index(inplace = True)
round_state_df.drop("index", axis = 1, inplace = True)
    
# Assuming round_state_df is your DataFrame
first_empty_row = get_first_empty_row(f"{output}/parsed.xlsx")

file_path = os.path.join(output, "parsed.csv")

# Assuming round_state_df is your DataFrame
if os.path.exists(file_path):
    # If the file exists, append without writing the header
    round_state_df.to_csv(file_path, mode='a', index=False, header=False)
else:
    # If the file does not exist, write with the header
    round_state_df.to_csv(file_path, mode='w', index=False, header=True)
    